import os
import sys
from tkinter import Button, Label, StringVar, Tk, filedialog, messagebox
import pandas as pd
from openpyxl.styles import Border, Font, PatternFill, Side
from PIL import Image, ImageTk

# --- Global Localization Dictionary ---
localization = {
    "language": "EN",
    "title": "CSV Data Processor",
    "selected file": "No file selected.",
    "instruction label": "SELECTED INPUT FILE:",
    "browse button": "BROWSE FILE",
    "execute button": "EXECUTE PROCESS",
    "warning title": "Warning",
    "warning message": "Please select a valid CSV file first before executing.",
    "selection dialogue": "Select CSV file",
    "completion title": "Success",
    "completion message": "Data successfully processed!\n\nSaved to: ",
    "error title": "Error",
    "error message": "An error occurred:\n\n"
}


def process_csv_to_excel(input_csv_path):
    global localization
    try:
        df = pd.read_csv(input_csv_path, sep=";", dtype=str)

        if "VEHICLE-NUMBER" in df.columns:
            df["VEHICLE-NUMBER"] = df["VEHICLE-NUMBER"].str[:17]

        if "ENGINE-NUMBER" in df.columns:
            df["ENGINE-NUMBER"] = df["ENGINE-NUMBER"].str[:14]

        if "DELIVERY-NUMBER" in df.columns:
            mid_part = df["DELIVERY-NUMBER"].str[4:6].fillna("")
            right_part = df["DELIVERY-NUMBER"].str[-2:].fillna("")
            lot_number_data = mid_part + right_part
            delivery_idx = df.columns.get_loc("DELIVERY-NUMBER")
            df.insert(delivery_idx + 1, "LOT-NUMBER", lot_number_data)

        if "CODES" in df.columns:
            unnamed_cols = [
                col for col in df.columns if str(col).startswith("Unnamed")
            ]

            def combine_row_codes(row):
                segments = []
                base_code = str(row["CODES"]).strip()
                if base_code and base_code.lower() != "nan":
                    segments.append(base_code)
                for col in unnamed_cols:
                    val = str(row[col]).strip()
                    if val and val.lower() != "nan":
                        segments.append(val)
                return "-".join(segments)

            df["COMBINED-CODE"] = df.apply(combine_row_codes, axis=1)
            codes_idx = df.columns.get_loc("CODES")
            combined_col_data = df.pop("COMBINED-CODE")
            df.insert(codes_idx, "COMBINED-CODE", combined_col_data)
            df.drop(columns=unnamed_cols, inplace=True)
            df.drop(columns=["CODES"], inplace=True)

        file_dir, file_name = os.path.split(input_csv_path)
        base_name = os.path.splitext(file_name)[0]
        output_excel_path = os.path.join(file_dir, base_name + "_processed.xlsx")

        with pd.ExcelWriter(output_excel_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Processed Data")
            worksheet = writer.sheets["Processed Data"]

            thin_border = Border(
                left=Side(style="thin", color="D3D3D3"),
                right=Side(style="thin", color="D3D3D3"),
                top=Side(style="thin", color="D3D3D3"),
                bottom=Side(style="thin", color="D3D3D3"),
            )
            header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            header_fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            highlight_fill = PatternFill(
                start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
            )
            highlight_cols = {
                "ORDER-NUMBER",
                "VEHICLE-NUMBER",
                "ENGINE-NUMBER",
                "DELIVERY-NUMBER",
                "LOT-NUMBER",
                "COMBINED-CODE",
            }

            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
                for cell in row:
                    cell.border = thin_border
                    if cell.row == 1:
                        cell.font = header_font
                        cell.fill = header_fill
                    if cell.row > 1:
                        current_header = worksheet.cell(
                            row=1, column=cell.column
                        ).value
                        if current_header in highlight_cols:
                            cell.fill = highlight_fill

            for col in worksheet.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                worksheet.column_dimensions[col_letter].width = max_len + 3

        messagebox.showinfo(
            f"{localization["completion title"]}",
            f"{localization["completion_msg"]}{localization["output_excel_path"]}",
        )
    except Exception as e:
        messagebox.showerror(f"{localization["error title"]}", f"{localization["error message"]}{e}")


class AppGUI:

    def __init__(self, root):
        global localization
        self.root = root

        # Dimensions
        self.window_width = 450
        self.window_height = 300
        self.root.geometry(f"{self.window_width}x{self.window_height}")
        self.root.resizable(False, False)

        self.bg_dark = "#121212"
        self.bg_panel = "#1E1E1E"
        self.fg_white = "#FFFFFF"
        self.fg_gray = "#A0A0A0"

        self.root.configure(bg=self.bg_dark)

        # Resolve asset path matching structure
        if getattr(sys, "frozen", False):
            self.base_path = sys._MEIPASS
        else:
            self.base_path = os.path.dirname(os.path.abspath(__file__))

        icon_filename = os.path.join(self.base_path, "icon.png")
        bg_filename = os.path.join(self.base_path, "bg1.png")
        gif_filename = os.path.join(self.base_path, "logo_t.gif")

        self.app_icon = None
        self.pil_bg_base = None
        self.gif_frames = []
        self.current_frame_idx = 0

        # Run string setup
        self.root.title(f"{localization["title"]}")
        self.file_path_var = StringVar()
        self.file_path_var.set(f"{localization["selected file"]}")

        # --- Load Title Bar Icon ---
        if os.path.exists(icon_filename):
            try:
                pil_icon = Image.open(icon_filename)
                self.app_icon = ImageTk.PhotoImage(pil_icon)
                self.root.iconphoto(False, self.app_icon)
            except Exception as e:
                print(f"Failed to load title bar icon: {e}")

        # --- Load Base Background Image ---
        if os.path.exists(bg_filename):
            try:
                self.pil_bg_base = Image.open(bg_filename).convert("RGBA")
                self.pil_bg_base = self.pil_bg_base.resize(
                    (self.window_width, self.window_height), Image.Resampling.LANCZOS
                )
            except Exception as e:
                print(f"Failed to load base background image: {e}")

        # --- Load and Process GIF Frames ---
        if os.path.exists(gif_filename):
            try:
                pil_gif = Image.open(gif_filename)
                idx = 0
                while True:
                    pil_gif.seek(idx)
                    frame_raw = pil_gif.copy().convert("RGBA")
                    frame_resized = frame_raw.resize((128, 128), Image.Resampling.LANCZOS)
                    self.gif_frames.append(frame_resized)
                    idx += 1
            except EOFError:
                pass
            except Exception as e:
                print(f"Error parsing animated GIF frames: {e}")

        # Master background wrapper label setup
        self.label_bg = Label(root, bd=0)
        self.label_bg.place(x=0, y=0, relwidth=1, relheight=1)

        if not self.pil_bg_base:
            self.label_bg.configure(bg=self.bg_dark)

        # 1. Text Labels
        self.label_title = Label(
            root,
            text=f"{localization["instruction label"]}",
            font=("Consolas", 10, "bold"),
            bg=self.bg_dark,
            fg=self.fg_white,
        )
        self.label_title.pack(pady=(140, 2))

        self.label_path = Label(
            root,
            textvariable=self.file_path_var,
            wraplength=400,
            fg=self.fg_gray,
            bg=self.bg_dark,
            font=("Consolas", 9),
        )
        self.label_path.pack(pady=(0, 5))

        # 2. Action Buttons
        self.btn_browse = Button(
            root,
            text=f"{localization["browse button"]}",
            width=18,
            font=("Consolas", 9, "bold"),
            bg=self.bg_panel,
            fg=self.fg_white,
            activebackground=self.fg_white,
            activeforeground=self.bg_dark,
            bd=1,
            relief="flat",
            command=self.browse_file,
        )
        self.btn_browse.pack(pady=3)

        self.btn_process = Button(
            root,
            text=f"{localization["execute button"]}",
            width=18,
            font=("Consolas", 9, "bold"),
            bg=self.fg_white,
            fg=self.bg_dark,
            activebackground=self.bg_panel,
            activeforeground=self.fg_white,
            bd=0,
            relief="flat",
            command=self.start_processing,
        )
        self.btn_process.pack(pady=3)

        self.btn_graphic = Button(
            root,
            bg="#1A1A1A",
            activebackground="#1A1A1A",
            command=self.change_language,
            bd=0,
            highlightthickness=0,
            relief="flat",
        )
        self.btn_graphic.pack(pady=3)

        self.update_button_flag()

        if self.pil_bg_base:
            self.label_title.configure(bg="#1A1A1A")
            self.label_path.configure(bg="#1A1A1A")

        self.animate_interface()

    def update_button_flag(self):
        global localization
        flag_file = os.path.join(self.base_path, f"{localization["language"]}.png")
        if os.path.exists(flag_file):
            try:
                pil_img = Image.open(flag_file).convert("RGBA")
                pil_resized = pil_img.resize((32, 32), Image.Resampling.LANCZOS)
                self.btn_image_compiled = ImageTk.PhotoImage(pil_resized)
                self.btn_graphic.configure(image=self.btn_image_compiled)
            except Exception as e:
                print(f"Failed to refresh flag image layout asset: {e}")

    def animate_interface(self):
        if self.pil_bg_base and self.gif_frames:
            dynamic_bg = self.pil_bg_base.copy()
            current_frame = self.gif_frames[self.current_frame_idx]
            dynamic_bg.alpha_composite(current_frame, dest=(161, 5))
            self.tk_rendered_bg = ImageTk.PhotoImage(dynamic_bg)
            self.label_bg.configure(image=self.tk_rendered_bg)
            self.current_frame_idx = (self.current_frame_idx + 1) % len(self.gif_frames)
            self.root.after(50, self.animate_interface)

        elif not self.pil_bg_base and self.gif_frames:
            solid_bg = Image.new("RGBA", (self.window_width, self.window_height), self.bg_dark)
            current_frame = self.gif_frames[self.current_frame_idx]
            solid_bg.alpha_composite(current_frame, dest=(161, 5))
            self.tk_rendered_bg = ImageTk.PhotoImage(solid_bg)
            self.label_bg.configure(image=self.tk_rendered_bg)
            self.current_frame_idx = (self.current_frame_idx + 1) % len(self.gif_frames)
            self.root.after(50, self.animate_interface)

    def browse_file(self):
        global localization
        chosen = filedialog.askopenfilename(
            title=f"{localization["selection dialogue"]}",
            filetypes=[("CSV files", "*.csv"), ("all files", "*.*")],
        )
        if chosen:
            self.file_path_var.set(chosen)

    def start_processing(self):
        global localization
        current_path = self.file_path_var.get()
        # Handles comparison across either translation variant context
        if current_path in ["No file selected.", "Chưa chọn dữ liệu.", ""] or not current_path:
            messagebox.showwarning(f"{localization["warning title"]}", f"{localization["warning message"]}")
            return
        process_csv_to_excel(current_path)

    def change_language(self):
        global localization
        if localization["language"] == "VI":
            localization["language"] = "EN"
            localization["title"] = "CSV Data Processor"
            localization["selected file"] = "No file selected."
            localization["instruction label"] = "SELECTED INPUT FILE:"
            localization["browse button"] = "BROWSE FILE"
            localization["execute button"] = "EXECUTE PROCESS"
            localization["warning title"] = "Warning"
            localization["warning message"] = "Please select a valid CSV file first before executing."
            localization["selection dialogue"] = "Select CSV file"
            localization["completion title"] = "Success"
            localization["completion message"] = "Data successfully processed!\n\nSaved to: "
            localization["error title"] = "Error"
            localization["error message"] = "An error occurred:\n\n"
        elif localization["language"] == "EN":
            localization["language"] = "VI"
            localization["title"] = "Làm sạch dữ liệu CSV cho production order"
            localization["selected file"] = "Chưa chọn dữ liệu."
            localization["instruction label"] = "DỮ LIỆU ĐẦU VÀO:"
            localization["browse button"] = "CHỌN TỆP DỮ LIỆU"
            localization["execute button"] = "BẮT ĐẦU"
            localization["warning title"] = "Cảnh báo"
            localization["warning message"] = "Xin vui lòng chọn tệp tin CSV phù hợp trước khi nhấn BẮT ĐẦU."
            localization["selection dialogue"] = "Chọn tệp tin CSV"
            localization["completion title"] = "Thành công"
            localization["completion message"] = "Toàn bộ dữ liệu đã được xử lý thành công!\n\nTệp tin được lưu vào: "
            localization["error title"] = "Lỗi"
            localization["error message"] = "Lỗi đã xuất hiện như sau:\n\n"

        # 2. DYNAMICALLY RE-CONFIGURE WIDGETS ON SCREEN
        self.root.title(f"{localization["title"]}")
        self.label_title.configure(text=f"{localization["instruction label"]}")
        self.btn_browse.configure(text=f"{localization["browse button"]}")
        self.btn_process.configure(text=f"{localization["execute button"]}")

        # Only reset path display fallback string if no file has been browsed yet
        if self.file_path_var.get() in ["No file selected.", "Chưa chọn dữ liệu."]:
            self.file_path_var.set(f"{localization["selected file"]}")

        # Update the active flag image to show the new option
        self.update_button_flag()


def run_gui():
    root = Tk()
    app = AppGUI(root)
    root.mainloop()


if __name__ == "__main__":
    run_gui()